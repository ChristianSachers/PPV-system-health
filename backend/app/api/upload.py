"""
XLSX Upload API Router - Campaign Data Processing Pipeline

This router implements the critical XLSX upload endpoint that coordinates:
1. File validation and safety checks
2. DataConverter: European number format conversion
3. RuntimeParser: Date range and ASAP format parsing
4. CampaignClassifier: Campaign vs Deal determination
5. Database persistence with error handling

Educational Focus: Shows how to orchestrate multiple services in a FastAPI endpoint
while maintaining proper error handling and business logic separation.

API Contract (from test specifications):
- POST /api/v1/campaigns/upload
- Accept multipart/form-data with XLSX file
- Return processing results with counts and validation errors
- Support partial success scenarios (some campaigns succeed, others fail)
"""

import io
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
import json

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

# Import our database and models
from ..database import get_db
from ..models.campaign import Campaign, UploadSession

# Import Phase 2 services
from ..services.data_conversion import DataConverter, ConversionError
from ..services.runtime_parser import RuntimeParser, RuntimeParseError
from ..services.campaign_classifier import CampaignClassifier

# Import XLSX processing (we'll need openpyxl or similar)
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl is required for XLSX processing. Install with: pip install openpyxl")

logger = logging.getLogger(__name__)

router = APIRouter()


class XLSXProcessor:
    """
    Service class for processing XLSX files into campaign data.

    Coordinates all Phase 2 services to transform raw XLSX data into
    validated Campaign model instances ready for database persistence.
    """

    def __init__(self):
        self.data_converter = DataConverter()
        self.runtime_parser = RuntimeParser()
        self.campaign_classifier = CampaignClassifier()

    def process_xlsx_file(self, file_content: io.BytesIO) -> Dict[str, Any]:
        """
        Process XLSX file content into campaign data.

        Args:
            file_content: XLSX file content as BytesIO

        Returns:
            Dict containing:
            - campaigns: List of successfully processed campaigns
            - errors: List of processing errors with row details
            - summary: Processing statistics
        """
        try:
            # Load XLSX workbook
            workbook = openpyxl.load_workbook(file_content, data_only=True)
            worksheet = workbook.active

            # Extract campaign data from worksheet
            campaigns = []
            errors = []
            row_number = 1

            # Get header row to understand column mapping
            headers = self._extract_headers(worksheet)
            logger.info(f"Detected XLSX headers: {headers}")

            # Process data rows (skip header)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_number += 1

                try:
                    # Convert row to campaign data
                    campaign_data = self._process_row(row, headers, row_number)

                    if campaign_data:
                        campaigns.append(campaign_data)

                except Exception as e:
                    error_detail = {
                        "row": row_number,
                        "error": str(e),
                        "data": [str(cell) for cell in row if cell is not None][:5]  # First 5 columns for context
                    }
                    errors.append(error_detail)
                    logger.warning(f"Row {row_number} processing failed: {e}")

            # Generate processing summary
            summary = {
                "total_rows": row_number - 1,  # Exclude header
                "successful_campaigns": len(campaigns),
                "failed_campaigns": len(errors),
                "success_rate": (len(campaigns) / (row_number - 1)) * 100 if row_number > 1 else 0
            }

            return {
                "campaigns": campaigns,
                "errors": errors,
                "summary": summary
            }

        except Exception as e:
            logger.error(f"XLSX processing failed: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"XLSX file processing failed: {e}"
            )

    def _extract_headers(self, worksheet: Worksheet) -> Dict[str, int]:
        """
        Extract column headers and map to expected field names.

        Args:
            worksheet: openpyxl worksheet

        Returns:
            Dict mapping field names to column indices
        """
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # Map header names to column indices
        headers = {}
        for idx, header in enumerate(header_row):
            if header:
                header_clean = str(header).strip().lower()

                # Map common header variations to our model fields
                if "id" in header_clean or "uuid" in header_clean:
                    headers["id"] = idx
                elif "deal/campaign" in header_clean and "name" in header_clean:
                    headers["name"] = idx
                elif "runtime" in header_clean:
                    headers["runtime"] = idx
                elif "impression" in header_clean and "goal" in header_clean:
                    headers["impression_goal"] = idx
                elif "budget" in header_clean:
                    headers["budget_eur"] = idx
                elif "cpm" in header_clean:
                    headers["cpm_eur"] = idx
                elif "buyer" in header_clean:
                    headers["buyer"] = idx

        return headers

    def _process_row(self, row: tuple, headers: Dict[str, int], row_number: int) -> Optional[Dict[str, Any]]:
        """
        Process a single row into campaign data.

        Args:
            row: Row data tuple
            headers: Column mapping
            row_number: Row number for error reporting

        Returns:
            Campaign data dictionary or None if row is empty
        """
        # Skip empty rows
        if not any(cell for cell in row):
            return None

        # Extract raw values using header mapping
        raw_data = {}
        for field, col_idx in headers.items():
            if col_idx < len(row):
                raw_data[field] = row[col_idx]

        # Validate required fields
        required_fields = ["id", "name", "runtime", "impression_goal", "budget_eur", "cpm_eur", "buyer"]
        missing_fields = [field for field in required_fields if field not in raw_data or raw_data[field] is None]

        if missing_fields:
            raise ValueError(f"Missing required fields: {missing_fields}")

        # Phase 2 Service Orchestration:
        campaign_data = {}

        try:
            # 1. UUID preservation (no conversion needed)
            campaign_data["id"] = str(raw_data["id"]).strip()

            # 2. Basic string fields
            campaign_data["name"] = str(raw_data["name"]).strip()
            campaign_data["buyer"] = str(raw_data["buyer"]).strip()

            # 3. DataConverter: European number format conversion
            campaign_data["impression_goal"] = self.data_converter.convert_impression_goal(str(raw_data["impression_goal"]))
            campaign_data["budget_eur"] = self.data_converter.convert_european_decimal(str(raw_data["budget_eur"]))
            campaign_data["cpm_eur"] = self.data_converter.convert_european_decimal(str(raw_data["cpm_eur"]))

            # 4. RuntimeParser: Date parsing and validation
            runtime_str = str(raw_data["runtime"]).strip()
            campaign_data["runtime"] = runtime_str

            # Parse runtime to extract dates (RuntimeParser will validate format)
            runtime_result = self.runtime_parser.parse_runtime(runtime_str)

            # Map parsed dates to model fields
            campaign_data["runtime_start"] = runtime_result.get("start_date")
            campaign_data["runtime_end"] = runtime_result.get("end_date")

            # 5. CampaignClassifier: Campaign vs Deal classification (implicit in model)
            # The Campaign model will handle classification via the buyer field

            return campaign_data

        except (ConversionError, RuntimeParseError, ValueError) as e:
            raise ValueError(f"Data conversion failed: {e}")


@router.post("/campaigns/upload", status_code=status.HTTP_201_CREATED)
async def upload_campaigns(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
) -> Dict[str, Any]:
    """
    Upload and process XLSX campaign file.

    This endpoint orchestrates the complete XLSX processing pipeline:
    1. File validation (format, size, content)
    2. XLSX parsing and data extraction
    3. European number format conversion
    4. Runtime parsing and date validation
    5. Campaign vs Deal classification
    6. Database persistence with error handling

    Args:
        file: Uploaded XLSX file
        db: Database session

    Returns:
        Processing results with success/error counts and campaign IDs

    Raises:
        HTTPException: For various validation and processing errors
    """
    logger.info(f"Processing XLSX upload: {file.filename}")

    # 1. File validation
    if not file.filename.endswith(('.xlsx', '.XLSX')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only XLSX files are supported. Please upload a valid Excel file."
        )

    if file.size and file.size > 50 * 1024 * 1024:  # 50MB limit
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="File size exceeds 50MB limit. Please upload a smaller file."
        )

    # 2. Create upload session for tracking
    upload_session = UploadSession(
        filename=file.filename,
        file_size=file.size or 0
    )
    upload_session.mark_processing_started()

    try:
        db.add(upload_session)
        db.commit()
        db.refresh(upload_session)

        # 3. Process XLSX file
        file_content = await file.read()
        file_buffer = io.BytesIO(file_content)

        processor = XLSXProcessor()
        processing_result = processor.process_xlsx_file(file_buffer)

        # 4. Persist campaigns to database
        campaign_ids = []
        persistence_errors = []

        for campaign_data in processing_result["campaigns"]:
            try:
                # Create Campaign model instance
                campaign = Campaign(**campaign_data)

                # Add to database
                db.add(campaign)
                db.commit()
                db.refresh(campaign)

                campaign_ids.append(campaign.id)
                logger.info(f"Successfully saved campaign: {campaign.id}")

            except IntegrityError as e:
                db.rollback()
                error_detail = {
                    "campaign_id": campaign_data.get("id", "unknown"),
                    "error": "Duplicate campaign ID or constraint violation",
                    "details": str(e)
                }
                persistence_errors.append(error_detail)
                logger.warning(f"Campaign persistence failed: {e}")

            except Exception as e:
                db.rollback()
                error_detail = {
                    "campaign_id": campaign_data.get("id", "unknown"),
                    "error": f"Database error: {e}",
                    "details": str(e)
                }
                persistence_errors.append(error_detail)
                logger.error(f"Unexpected campaign persistence error: {e}")

        # 5. Update upload session with results
        total_campaigns = len(processing_result["campaigns"])
        successful_campaigns = len(campaign_ids)
        failed_campaigns = len(processing_result["errors"]) + len(persistence_errors)

        upload_session.mark_completed(
            successful=successful_campaigns,
            failed=failed_campaigns,
            total=total_campaigns + failed_campaigns
        )

        # Store errors as JSON for detailed reporting
        all_errors = processing_result["errors"] + persistence_errors
        if all_errors:
            upload_session.validation_errors = json.dumps(all_errors)

        db.commit()

        # 6. Prepare response
        response_data = {
            "processed_count": successful_campaigns,
            "failed_count": failed_campaigns,
            "campaign_ids": campaign_ids,
            "upload_session_id": upload_session.id,
            "summary": {
                **processing_result["summary"],
                "persistence_errors": len(persistence_errors)
            }
        }

        # Include error details if there were failures
        if all_errors:
            response_data["errors"] = all_errors[:10]  # Limit to first 10 errors
            response_data["total_errors"] = len(all_errors)

        # Return appropriate status code
        if failed_campaigns > 0 and successful_campaigns > 0:
            # Partial success
            return JSONResponse(
                status_code=status.HTTP_207_MULTI_STATUS,
                content=response_data
            )
        elif successful_campaigns == 0:
            # Complete failure
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No campaigns could be processed successfully",
                headers={"X-Processing-Errors": str(len(all_errors))}
            )
        else:
            # Complete success
            return response_data

    except HTTPException:
        # Re-raise HTTP exceptions
        upload_session.mark_failed("HTTP validation error")
        db.commit()
        raise

    except Exception as e:
        # Handle unexpected errors
        logger.error(f"XLSX upload processing failed: {e}", exc_info=True)
        upload_session.mark_failed(f"Processing error: {e}")
        db.commit()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Campaign upload processing failed: {e}"
        )


@router.get("/upload/status/{session_id}")
async def get_upload_status(
    session_id: str,
    db: Session = Depends(get_db)
) -> Dict[str, Any]:
    """
    Get upload session status and processing details.

    Args:
        session_id: Upload session ID
        db: Database session

    Returns:
        Upload session details with processing status
    """
    upload_session = db.query(UploadSession).filter(UploadSession.id == session_id).first()

    if not upload_session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Upload session {session_id} not found"
        )

    response_data = {
        "session_id": upload_session.id,
        "filename": upload_session.filename,
        "status": upload_session.status,
        "upload_timestamp": upload_session.upload_timestamp.isoformat(),
        "processing_started_at": upload_session.processing_started_at.isoformat() if upload_session.processing_started_at else None,
        "processing_completed_at": upload_session.processing_completed_at.isoformat() if upload_session.processing_completed_at else None,
        "results": {
            "total_rows_processed": upload_session.total_rows_processed,
            "successful_campaigns": upload_session.successful_campaigns,
            "failed_campaigns": upload_session.failed_campaigns
        }
    }

    # Include validation errors if present
    if upload_session.validation_errors:
        try:
            response_data["validation_errors"] = json.loads(upload_session.validation_errors)
        except json.JSONDecodeError:
            response_data["validation_errors"] = upload_session.validation_errors

    return response_data